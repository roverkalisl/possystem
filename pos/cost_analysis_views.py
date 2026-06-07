"""
Project Cost Analysis Views
Handles budget management, cost analysis reports, and dashboards
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Sum, Q, F
from django.utils import timezone
from decimal import Decimal
import json
from datetime import datetime
from io import BytesIO
import csv

from .models import (
    Project, ProjectBudget, ProjectBudgetLine, ProjectExpense,
    ProjectPettyCashExpense, GRNItem, GLMaster
)
from .forms import BudgetUploadForm, ProjectBudgetForm, ProjectBudgetLineFormSet
from .cost_analysis import ProjectCostAnalyzer


def has_cost_analysis_permission(user):
    """Check if user has permission to access cost analysis"""
    return user.is_superuser or user.is_staff or user.groups.filter(
        name__in=['Finance Manager', 'Project Manager', 'Owner']
    ).exists()


@login_required
@require_http_methods(["GET"])
def project_cost_analysis_list(request):
    """List all projects with cost analysis available"""
    if not has_cost_analysis_permission(request.user):
        messages.error(request, 'You do not have permission to access cost analysis')
        return redirect('dashboard')
    
    projects = Project.objects.filter(is_active=True).prefetch_related('budget')
    
    context = {
        'projects': projects,
        'page_title': 'Project Cost Analysis'
    }
    return render(request, 'pos/cost_analysis_list.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def budget_upload(request):
    """Upload project budget from Excel file"""
    if not has_cost_analysis_permission(request.user):
        messages.error(request, 'You do not have permission to upload budgets')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = BudgetUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                project = form.cleaned_data['project']
                excel_file = request.FILES['budget_file']
                replace_existing = form.cleaned_data.get('replace_existing', False)
                create_missing_gl = form.cleaned_data.get('create_missing_gl', False)
                import_multiple_projects = form.cleaned_data.get('import_multiple_projects', False)

                # Process Excel file
                success_count, error_messages, projects_updated = process_budget_excel(
                    project, excel_file, replace_existing, create_missing_gl, import_multiple_projects, request.user
                )
                
                if error_messages:
                    for error in error_messages:
                        messages.warning(request, error)
                
                messages.success(
                    request,
                    f'Budget imported successfully. {success_count} GL accounts added.'
                )
                # Redirect: if multiple projects were updated go to list, else to project detail
                if import_multiple_projects:
                    return redirect('project_cost_analysis_list')
                else:
                    return redirect('project_cost_analysis_detail', project_id=project.id)
            
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = BudgetUploadForm()
    
    context = {
        'form': form,
        'page_title': 'Upload Budget'
    }
    return render(request, 'pos/budget_upload.html', context)


@login_required
@require_http_methods(["GET"])
def project_cost_analysis_detail(request, project_id):
    """Detailed cost analysis for a specific project"""
    if not has_cost_analysis_permission(request.user):
        messages.error(request, 'You do not have permission to view cost analysis')
        return redirect('dashboard')
    
    project = get_object_or_404(Project, id=project_id, is_active=True)
    analyzer = ProjectCostAnalyzer(project)
    
    # Get summary data
    summary = analyzer.get_cost_summary()
    
    # Get date range filter
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    context = {
        'project': project,
        'summary': summary,
        'page_title': f'Cost Analysis - {project.project_id}',
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'pos/cost_analysis_detail.html', context)


@login_required
@require_http_methods(["GET"])
def cost_analysis_by_gl_group(request):
    """GL Group-wise cost analysis report"""
    if not has_cost_analysis_permission(request.user):
        messages.error(request, 'You do not have permission to view reports')
        return redirect('dashboard')
    
    # Get filters
    project_id = request.GET.get('project_id')
    gl_group = request.GET.get('gl_group')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    projects = Project.objects.filter(is_active=True)
    summary_data = {}
    
    if project_id:
        projects = projects.filter(id=project_id)
    
    for project in projects:
        analyzer = ProjectCostAnalyzer(project)
        if gl_group:
            # Filter by specific GL group
            full_summary = analyzer.get_cost_summary()
            if gl_group in full_summary['by_group']:
                summary_data[project.id] = {
                    'project': project,
                    'data': {gl_group: full_summary['by_group'][gl_group]},
                    'total': full_summary['by_group'][gl_group]
                }
        else:
            # All GL groups
            summary = analyzer.get_cost_summary()
            summary_data[project.id] = {
                'project': project,
                'data': summary['by_group'],
                'total': summary['totals']
            }
    
    context = {
        'summary_data': summary_data,
        'projects': projects,
        'gl_groups': [
            'Direct Material Cost', 'Direct Labour Cost', 'Plant & Equipment Cost',
            'Subcontractor Cost', 'Third Party Service Cost', 'Transportation & Logistics',
            'Project-Specific Expenses', 'Variations / Cost Adjustments',
            'Retail Operating Expenses', 'Overhead Expenses', 'Selling & Marketing',
            'General Expenses', 'Other Expenses'
        ],
        'page_title': 'GL Group Cost Analysis',
        'selected_project': project_id,
        'selected_group': gl_group,
    }
    return render(request, 'pos/cost_analysis_gl_group.html', context)


@login_required
@require_http_methods(["GET"])
def gl_master_group_report(request):
    """Report: list GLMaster grouped by GL Group"""
    if not has_cost_analysis_permission(request.user):
        messages.error(request, 'You do not have permission to view reports')
        return redirect('dashboard')

    gls = GLMaster.objects.filter(is_active=True).order_by('gl_code')
    groups = {}
    for gl in gls:
        group_name = ProjectBudgetLine.get_gl_group(gl.gl_code)
        groups.setdefault(group_name, []).append(gl)

    context = {
        'groups': groups,
        'page_title': 'GL Master Group Report'
    }
    return render(request, 'pos/gl_master_group_report.html', context)


@login_required
@require_http_methods(["GET"])
def transaction_details(request, project_id):
    """Show detailed transactions for a GL code or GL group"""
    if not has_cost_analysis_permission(request.user):
        messages.error(request, 'You do not have permission to view transactions')
        return redirect('dashboard')
    
    project = get_object_or_404(Project, id=project_id, is_active=True)
    gl_code = request.GET.get('gl_code')
    gl_group = request.GET.get('gl_group')
    
    analyzer = ProjectCostAnalyzer(project)
    transactions = analyzer.get_transaction_details(gl_code=gl_code, gl_group=gl_group)
    
    context = {
        'project': project,
        'transactions': transactions,
        'gl_code': gl_code,
        'gl_group': gl_group,
        'page_title': f'Transaction Details - {project.project_id}',
    }
    return render(request, 'pos/transaction_details.html', context)


@login_required
@require_http_methods(["GET"])
def export_cost_analysis(request, project_id):
    """Export cost analysis to Excel"""
    try:
        import openpyxl
    except ImportError:
        messages.error(request, 'openpyxl package is not installed. Please contact administrator.')
        return redirect('project_cost_analysis_detail', project_id=project_id)
    
    if not has_cost_analysis_permission(request.user):
        messages.error(request, 'You do not have permission to export')
        return redirect('dashboard')
    
    project = get_object_or_404(Project, id=project_id, is_active=True)
    analyzer = ProjectCostAnalyzer(project)
    summary = analyzer.get_cost_summary()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cost Analysis'
    
    # Header
    ws['A1'] = 'Project Cost Analysis Report'
    ws['A2'] = f'Project: {project.project_id} - {project.project_name}'
    ws['A3'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
    
    # GL-wise Summary
    row = 5
    ws[f'A{row}'] = 'GL Code'
    ws[f'B{row}'] = 'GL Name'
    ws[f'C{row}'] = 'Budget'
    ws[f'D{row}'] = 'Actual'
    ws[f'E{row}'] = 'Variance'
    ws[f'F{row}'] = 'Utilization %'
    
    row += 1
    for gl_code in sorted(summary['by_gl'].keys()):
        data = summary['by_gl'][gl_code]
        ws[f'A{row}'] = gl_code
        ws[f'B{row}'] = data['gl_name']
        ws[f'C{row}'] = data['budget']
        ws[f'D{row}'] = data['actual']
        ws[f'E{row}'] = data['variance']
        ws[f'F{row}'] = round(data['utilization_percent'], 2)
        row += 1
    
    # Totals
    row += 1
    ws[f'A{row}'] = 'TOTAL'
    ws[f'C{row}'] = summary['totals']['budget']
    ws[f'D{row}'] = summary['totals']['actual']
    ws[f'E{row}'] = summary['totals']['variance']
    ws[f'F{row}'] = round(summary['totals']['utilization_percent'], 2)
    
    # GL Group Summary
    row += 3
    ws[f'A{row}'] = 'GL Group Summary'
    row += 1
    ws[f'A{row}'] = 'GL Group'
    ws[f'B{row}'] = 'Budget'
    ws[f'C{row}'] = 'Actual'
    ws[f'D{row}'] = 'Variance'
    ws[f'E{row}'] = 'Utilization %'
    
    row += 1
    for group_name in sorted(summary['by_group'].keys()):
        data = summary['by_group'][group_name]
        ws[f'A{row}'] = group_name
        ws[f'B{row}'] = data['budget']
        ws[f'C{row}'] = data['actual']
        ws[f'D{row}'] = data['variance']
        ws[f'E{row}'] = round(data.get('utilization_percent', 0), 2)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = \
        f'attachment; filename="Cost_Analysis_{project.project_id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@require_http_methods(["GET"])
def export_cost_analysis_csv(request, project_id):
    """Export cost analysis to CSV"""
    if not has_cost_analysis_permission(request.user):
        messages.error(request, 'You do not have permission to export')
        return redirect('dashboard')
    
    project = get_object_or_404(Project, id=project_id, is_active=True)
    analyzer = ProjectCostAnalyzer(project)
    summary = analyzer.get_cost_summary()
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = \
        f'attachment; filename="Cost_Analysis_{project.project_id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Header
    writer.writerow(['Project Cost Analysis Report'])
    writer.writerow([f'Project: {project.project_id} - {project.project_name}'])
    writer.writerow([f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    writer.writerow([])
    
    # GL-wise Summary
    writer.writerow(['GL Code', 'GL Name', 'Budget', 'Actual', 'Variance', 'Utilization %'])
    for gl_code in sorted(summary['by_gl'].keys()):
        data = summary['by_gl'][gl_code]
        writer.writerow([
            gl_code,
            data['gl_name'],
            data['budget'],
            data['actual'],
            data['variance'],
            round(data['utilization_percent'], 2)
        ])
    
    # Totals
    writer.writerow([])
    writer.writerow([
        'TOTAL', '', summary['totals']['budget'], summary['totals']['actual'],
        summary['totals']['variance'], round(summary['totals']['utilization_percent'], 2)
    ])
    
    # GL Group Summary
    writer.writerow([])
    writer.writerow(['GL Group Summary'])
    writer.writerow(['GL Group', 'Budget', 'Actual', 'Variance', 'Utilization %'])
    for group_name in sorted(summary['by_group'].keys()):
        data = summary['by_group'][group_name]
        writer.writerow([
            group_name,
            data['budget'],
            data['actual'],
            data['variance'],
            round(data.get('utilization_percent', 0), 2)
        ])
    
    return response


@login_required
@require_http_methods(["GET"])
def export_gl_group_project_report(request):
    """Export a CSV report: Project x GL Group with Budget, Actual, Variance, Utilization % and Status"""
    if not has_cost_analysis_permission(request.user):
        messages.error(request, 'You do not have permission to export')
        return redirect('dashboard')

    projects = Project.objects.filter(is_active=True)

    # Prepare CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="GL_Group_Project_Report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    writer = csv.writer(response)

    # Header
    writer.writerow(['Project', 'GL Group (GL Name)', 'Budget', 'Actual Cost', 'Variance', 'Utilization %', 'Status'])

    for project in projects:
        analyzer = ProjectCostAnalyzer(project)
        summary = analyzer.get_cost_summary()
        by_group = summary.get('by_group', {})

        for group_name, data in by_group.items():
            # group-level budget and actual
            budget_amt = data.get('budget', 0)
            actual_amt = data.get('actual', 0)
            variance = data.get('variance', 0)
            utilization = round(data.get('utilization_percent', 0), 2)

            # Try to include a representative GL name (first GL in that group for project)
            gl_name = ''
            try:
                # find any budget line in project with that group
                line = ProjectBudgetLine.objects.filter(budget__project=project).select_related('gl_account').all().first()
                if line:
                    gl_name = line.gl_account.gl_name
            except Exception:
                gl_name = ''

            status = project.budget.get_status_display() if hasattr(project, 'budget') and project.budget else 'No Budget'

            writer.writerow([
                project.project_id,
                f"{group_name} ({gl_name})",
                budget_amt,
                actual_amt,
                variance,
                utilization,
                status
            ])

    return response


def process_budget_excel(project, excel_file, replace_existing=False, create_missing_gl=False, import_multiple_projects=False, user=None):
    """
    Process budget Excel file and create/update budget records
    
    Expected Excel structure:
    GL Code | GL Name | Budget Amount
    """
    try:
        import openpyxl
    except ImportError:
        return 0, ['openpyxl package is not installed. Please contact administrator.']
    
    success_count = 0
    error_messages = []
    projects_updated = set()
    seen_keys = set()  # (project_id, gl_code) to detect duplicates in file
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # Read header row to detect columns
        header = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header = [str(c).strip().lower() if c is not None else '' for c in row]
            break

        # helper to find column index by possible names
        def find_col(names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None

        project_col = find_col(['project', 'project_id', 'project number', 'project no', 'project_no'])
        gl_code_col = find_col(['gl code', 'gl_code', 'glcode']) or 0
        gl_name_col = find_col(['gl name', 'gl_name', 'glname']) or 1
        amount_col = find_col(['budget amount', 'budget_amount', 'amount', 'budget']) or 2

        budgets_cache = {}  # project_id -> ProjectBudget instance


        # First pass: accumulate amounts per (project, gl_code)
        pending = {}  # key -> {'project': proj_obj, 'gl_name': str, 'amount': Decimal}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Determine project for this row
                if import_multiple_projects and project_col is not None:
                    proj_val = row[project_col]
                    if not proj_val:
                        error_messages.append(f'Row {row_idx}: Missing project identifier')
                        continue
                    proj_code = str(proj_val).strip()
                    try:
                        proj_obj = Project.objects.get(project_id=proj_code)
                    except Project.DoesNotExist:
                        error_messages.append(f'Row {row_idx}: Project {proj_code} not found')
                        continue
                else:
                    proj_obj = project

                # columns
                gl_code = row[gl_code_col] if len(row) > gl_code_col else None
                gl_name = row[gl_name_col] if len(row) > gl_name_col else None
                budget_amount = row[amount_col] if len(row) > amount_col else None

                if gl_code is None or budget_amount is None:
                    error_messages.append(f'Row {row_idx}: Missing GL Code or Budget Amount')
                    continue

                gl_code = str(gl_code).strip()
                gl_name = str(gl_name).strip() if gl_name else ''

                # parse amount
                try:
                    budget_amount_dec = Decimal(str(budget_amount))
                except Exception:
                    error_messages.append(f'Row {row_idx}: Invalid budget amount format')
                    continue

                key = (proj_obj.id, gl_code)
                if key in pending:
                    pending[key]['amount'] += budget_amount_dec
                else:
                    pending[key] = {
                        'project': proj_obj,
                        'gl_name': gl_name,
                        'amount': budget_amount_dec
                    }

            except Exception as e:
                error_messages.append(f'Row {row_idx}: {str(e)}')

        # Second pass: write aggregated data to DB
        # Track which projects we've cleared (for replace_existing)
        cleared_projects = set()
        from .models import GLCreationLog

        for (proj_id, gl_code), info in pending.items():
            proj_obj = info['project']
            gl_name = info['gl_name']
            amt = info['amount']

            # Ensure GL exists
            try:
                gl_account = GLMaster.objects.get(gl_code=str(gl_code))
                created_gl = False
            except GLMaster.DoesNotExist:
                if create_missing_gl:
                    parent_group = ProjectBudgetLine.get_gl_group(str(gl_code)) if hasattr(ProjectBudgetLine, 'get_gl_group') else ''
                    gl_account = GLMaster.objects.create(
                        gl_code=str(gl_code),
                        gl_name=gl_name or str(gl_code),
                        gl_type='expense',
                        parent_group=parent_group
                    )
                    created_gl = True
                    # log creation
                    try:
                        GLCreationLog.objects.create(
                            gl=gl_account,
                            created_by=user if user and hasattr(user, 'id') else None,
                            source='budget_import'
                        )
                    except Exception:
                        # don't fail import if logging fails
                        pass
                else:
                    error_messages.append(f'Project {proj_obj.project_id}: GL Code {gl_code} not found')
                    continue

            # get or create budget for project
            if proj_obj.id not in budgets_cache:
                b, created = ProjectBudget.objects.get_or_create(
                    project=proj_obj,
                    defaults={'status': 'active'}
                )
                budgets_cache[proj_obj.id] = b

            # clear if requested (do once per project)
            budget_obj = budgets_cache[proj_obj.id]
            if replace_existing and proj_obj.id not in cleared_projects:
                budget_obj.lines.all().delete()
                cleared_projects.add(proj_obj.id)

            # Create or update budget line with aggregated amount
            ProjectBudgetLine.objects.update_or_create(
                budget=budget_obj,
                gl_account=gl_account,
                defaults={'budget_amount': amt}
            )

            projects_updated.add(proj_obj.id)
            success_count += 1

        # Recalculate totals for updated budgets
        for pid in projects_updated:
            b = budgets_cache.get(pid) or ProjectBudget.objects.get(project_id=pid)
            b.recalculate_total()

    except Exception as e:
        error_messages.append(f'File processing error: {str(e)}')

    return success_count, error_messages, list(projects_updated)


@login_required
@require_http_methods(["GET"])
def project_dashboard_cost_summary(request, project_id):
    """Get cost summary for project dashboard (API endpoint)"""
    if not has_cost_analysis_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        project = Project.objects.get(id=project_id, is_active=True)
        analyzer = ProjectCostAnalyzer(project)
        summary = analyzer.get_cost_summary()
        
        return JsonResponse({
            'status': 'success',
            'data': {
                'budget': summary['totals']['budget'],
                'actual': summary['totals']['actual'],
                'variance': summary['totals']['variance'],
                'utilization_percent': round(summary['totals']['utilization_percent'], 2),
            }
        })
    except Project.DoesNotExist:
        return JsonResponse({'error': 'Project not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
